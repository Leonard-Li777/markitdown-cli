import os
import io
import fitz  # PyMuPDF
import openpyxl
from ._markitdown import MarkItDown, StreamInfo, DocumentConverterResult
from ._page_range import parse_pages, resolve

def route_document(
    file_path: str,
    file_bytes: bytes,
    extension: str,
    enable_ocr: bool = False,
    pages_spec_str: str = None,
    **kwargs
) -> str:
    """
    Main router for document processing.
    Handles size boundaries, in-memory PDF splitting, and lightweight office parsing.
    """
    file_size_mb = len(file_bytes) / (1024 * 1024)
    ext = (extension or "").lower()

    # 1. Handle Office formats to PDF conversion for OCR
    office_exts = {".docx", ".doc", ".pptx", ".ppt", ".xlsx", ".xls", ".odt", ".odp", ".ods"}
    if ext in office_exts:
        if enable_ocr:
            try:
                from ._pdf_output import office_to_pdf, PdfConversionError
                pages_spec = parse_pages(pages_spec_str) if pages_spec_str else None
                # office_to_pdf automatically converts office using LibreOffice and filters pages if requested
                pdf_bytes = office_to_pdf(file_path, pages_spec=pages_spec)
                # Recursively route the generated PDF bytes (the pages are already filtered, so clear pages_spec_str)
                return route_document(
                    file_path=file_path,
                    file_bytes=pdf_bytes,
                    extension=".pdf",
                    enable_ocr=enable_ocr,
                    pages_spec_str=None,
                    **kwargs
                )
            except Exception as e:
                # Graceful fallback: if LibreOffice is not installed or conversion fails,
                # fall back to direct non-OCR extraction (still useful for text-based Office files)
                import warnings
                warnings.warn(
                    f"OCR conversion failed for Office file ({type(e).__name__}: {e}). "
                    f"Falling back to direct non-OCR text extraction.",
                    RuntimeWarning
                )
        elif ext == ".xlsx" and file_size_mb > 20:
            # For non-OCR large spreadsheets, use lightweight streaming to prevent memory blow-up
            return lightweight_xlsx_text_extract(file_bytes)

    # 2. Handle PDF routing based on size and OCR mode
    if ext == ".pdf":
        if not enable_ocr:
            if pages_spec_str is not None:
                # If specific pages are requested, always split in memory.
                # This guarantees that only the selected pages are processed,
                # and bypasses pdfminer bugs that ignore page selection.
                return pdf_memory_split_and_process(
                    file_bytes=file_bytes,
                    chunk_size=50,
                    pages_spec_str=pages_spec_str,
                    enable_ocr=False,
                    **kwargs
                )
            elif file_size_mb <= 100:
                return markitdown_direct_convert(
                    file_bytes=file_bytes,
                    extension=ext,
                    pages_spec_str=None,
                    enable_ocr=False,
                    **kwargs
                )
            else:
                # Split large PDF into 50-page chunks in memory
                return pdf_memory_split_and_process(
                    file_bytes=file_bytes,
                    chunk_size=50,
                    pages_spec_str=None,
                    enable_ocr=False,
                    **kwargs
                )
        else:
            if pages_spec_str is not None:
                # Always split in memory if pages are specified to control memory in OCR mode
                return pdf_memory_split_and_process(
                    file_bytes=file_bytes,
                    chunk_size=5,
                    pages_spec_str=pages_spec_str,
                    enable_ocr=True,
                    **kwargs
                )
            elif file_size_mb <= 5:
                return markitdown_direct_convert(
                    file_bytes=file_bytes,
                    extension=ext,
                    pages_spec_str=None,
                    enable_ocr=True,
                    **kwargs
                )
            else:
                # Split large OCR PDF into 5-page chunks in memory
                return pdf_memory_split_and_process(
                    file_bytes=file_bytes,
                    chunk_size=5,
                    pages_spec_str=None,
                    enable_ocr=True,
                    **kwargs
                )

    # 3. Fallback/Direct conversion
    return markitdown_direct_convert(
        file_bytes=file_bytes,
        extension=ext,
        pages_spec_str=pages_spec_str,
        enable_ocr=enable_ocr,
        **kwargs
    )

def markitdown_direct_convert(
    file_bytes: bytes,
    extension: str,
    pages_spec_str: str = None,
    enable_ocr: bool = False,
    **kwargs
) -> str:
    """
    Directly converts a document stream in memory using MarkItDown.
    """
    md_kwargs = {}
    if enable_ocr:
        md_kwargs["enable_plugins"] = True
        md_kwargs["use_tesseract"] = kwargs.get("ocr_engine", "tesseract") == "tesseract"
        if kwargs.get("ocr_engine", "tesseract") == "tesseract":
            if kwargs.get("tesseract_path"):
                md_kwargs["tesseract_path"] = kwargs.get("tesseract_path")
            md_kwargs["tesseract_lang"] = kwargs.get("tesseract_lang", "eng")
        elif kwargs.get("ocr_engine") == "llm":
            if kwargs.get("llm_client"):
                md_kwargs["llm_client"] = kwargs.get("llm_client")
            if kwargs.get("llm_model"):
                md_kwargs["llm_model"] = kwargs.get("llm_model")
    else:
        md_kwargs["enable_plugins"] = kwargs.get("enable_plugins", False)

    # Pass other standard kwargs
    for key in ["with_metadata", "metadata_only", "docintel_endpoint", "cu_endpoint", "cu_analyzer_id", "cu_file_types"]:
        if key in kwargs:
            md_kwargs[key] = kwargs[key]

    markitdown = MarkItDown(**md_kwargs)

    convert_kwargs = {}
    if pages_spec_str:
        convert_kwargs["pages"] = parse_pages(pages_spec_str)

    stream = io.BytesIO(file_bytes)
    result = markitdown.convert_stream(
        stream,
        stream_info=StreamInfo(extension=extension),
        **convert_kwargs
    )
    return result.markdown

def pdf_memory_split_and_process(
    file_bytes: bytes,
    chunk_size: int,
    pages_spec_str: str = None,
    enable_ocr: bool = False,
    **kwargs
) -> str:
    """
    Splits a PDF in memory using PyMuPDF (fitz) and processes each chunk independently.
    """
    try:
        doc = fitz.open("pdf", file_bytes)
    except Exception:
        doc = fitz.open(stream=file_bytes, filetype="pdf")

    total_pages = doc.page_count

    # Resolve actual page selection
    pages_spec = parse_pages(pages_spec_str) if pages_spec_str else None
    resolved_pages = resolve(pages_spec, total_pages)
    if resolved_pages is None:
        resolved_pages = list(range(1, total_pages + 1))
    else:
        resolved_pages = sorted(list(resolved_pages))

    results = []
    # Chunk the page indices
    for i in range(0, len(resolved_pages), chunk_size):
        chunk_page_nums = resolved_pages[i:i + chunk_size]

        # Filter page numbers to ensure they are within valid bounds
        valid_page_nums = [p for p in chunk_page_nums if 1 <= p <= total_pages]
        if not valid_page_nums:
            continue

        # Extract selected pages into a new in-memory PDF document
        chunk_doc = fitz.open()
        try:
            for page_num in valid_page_nums:
                chunk_doc.insert_pdf(doc, from_page=page_num - 1, to_page=page_num - 1)
            chunk_bytes = chunk_doc.tobytes()
        finally:
            chunk_doc.close()

        # Convert chunk bytes directly in memory.
        # Clear pages_spec_str because chunk_bytes contains exactly the pages we want.
        chunk_md = markitdown_direct_convert(
            file_bytes=chunk_bytes,
            extension=".pdf",
            pages_spec_str=None,
            enable_ocr=enable_ocr,
            **kwargs
        )
        
        # Post-process chunk markdown to restore original page numbers in OCR headers (e.g., "## Page X")
        if chunk_md and chunk_md.strip():
            import re
            for idx, original_page_num in enumerate(valid_page_nums, 1):
                chunk_md = re.sub(
                    rf"(^|\n)## Page {idx}(?=\n|$)",
                    rf"\1## Page {original_page_num}",
                    chunk_md
                )
            results.append(chunk_md.strip())

    doc.close()
    return "\n\n".join(results)

def lightweight_xlsx_text_extract(file_bytes: bytes) -> str:
    """
    Extracts text from Excel (.xlsx) files using openpyxl(read_only=True)
    to maintain a low memory footprint for large spreadsheets.
    """
    stream = io.BytesIO(file_bytes)
    wb = openpyxl.load_workbook(stream, read_only=True, data_only=True)
    md_content = []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        md_content.append(f"## {sheet_name}\n")

        is_first = True
        for row in sheet.iter_rows(values_only=True):
            # Skip completely empty rows
            if not any(val is not None for val in row):
                continue

            # Format cell values
            cells = [str(val).replace("\n", " ").strip() if val is not None else "" for val in row]
            cells = [c.replace("|", "\\|") for c in cells]

            row_str = "| " + " | ".join(cells) + " |"
            md_content.append(row_str)

            if is_first:
                sep_str = "| " + " | ".join(["---"] * len(cells)) + " |"
                md_content.append(sep_str)
                is_first = False

        md_content.append("")

    wb.close()
    return "\n".join(md_content).strip()
